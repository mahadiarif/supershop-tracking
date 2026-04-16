from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


class ReportService:
    """রিপোর্ট এক্সপোর্টের জন্য ছোট helper layer."""

    @staticmethod
    async def export_to_excel(data: list[dict[str, Any]], sheet_name: str = "Report") -> BytesIO:
        """Generic tabular data Excel এ লিখে, chart সহ BytesIO ফেরত দেয়."""
        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        output.seek(0)
        return output

    @staticmethod
    async def generate_daily_excel(report: dict[str, Any], target_date: date) -> BytesIO:
        return await ReportService._generate_chart_excel(
            rows=report.get("hourly_breakdown", []),
            sheet_name="Daily Report",
            title=f"Daily Report - {target_date.isoformat()}",
            chart_type="bar",
            x_field="hour",
            y_field="count",
        )

    @staticmethod
    async def generate_weekly_excel(report: dict[str, Any], start_date: date) -> BytesIO:
        return await ReportService._generate_chart_excel(
            rows=report.get("daily_breakdown", []),
            sheet_name="Weekly Report",
            title=f"Weekly Report - {start_date.isoformat()}",
            chart_type="line",
            x_field="date",
            y_field="value",
        )

    @staticmethod
    async def generate_monthly_excel(report: dict[str, Any], month: str) -> BytesIO:
        return await ReportService._generate_chart_excel(
            rows=report.get("daily_breakdown", []),
            sheet_name="Monthly Report",
            title=f"Monthly Report - {month}",
            chart_type="line",
            x_field="date",
            y_field="value",
        )

    @staticmethod
    async def generate_incidents_excel(rows: list[dict[str, Any]], start: datetime, end: datetime) -> BytesIO:
        severity_rows = []
        counts = {}
        for row in rows:
            key = row.get("severity", "unknown")
            counts[key] = counts.get(key, 0) + 1
        for key, value in counts.items():
            severity_rows.append({"severity": key, "count": value})

        return await ReportService._generate_chart_excel(
            rows=severity_rows,
            sheet_name="Incidents",
            title=f"Incidents - {start.date().isoformat()} to {end.date().isoformat()}",
            chart_type="pie",
            x_field="severity",
            y_field="count",
        )

    @staticmethod
    async def _generate_chart_excel(
        rows: list[dict[str, Any]],
        sheet_name: str,
        title: str,
        chart_type: str,
        x_field: str,
        y_field: str,
    ) -> BytesIO:
        """একটি simple Excel sheet + chart generate করে."""
        df = pd.DataFrame(rows if rows else [{x_field: "No data", y_field: 0}])
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        output.seek(0)
        workbook = load_workbook(output)
        worksheet = workbook[sheet_name]
        worksheet["A1"] = title

        if chart_type == "bar":
            chart = BarChart()
            chart.title = title
            chart.y_axis.title = "Count"
            chart.x_axis.title = x_field
            data = Reference(worksheet, min_col=2, min_row=1, max_row=worksheet.max_row)
            cats = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            worksheet.add_chart(chart, "D4")
        elif chart_type == "line":
            chart = LineChart()
            chart.title = title
            data = Reference(worksheet, min_col=2, min_row=1, max_row=worksheet.max_row)
            cats = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            worksheet.add_chart(chart, "D4")
        elif chart_type == "pie":
            chart = PieChart()
            chart.title = title
            data = Reference(worksheet, min_col=2, min_row=1, max_row=worksheet.max_row)
            labels = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            worksheet.add_chart(chart, "D4")

        final_output = BytesIO()
        workbook.save(final_output)
        final_output.seek(0)
        return final_output


report_service = ReportService()
